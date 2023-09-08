"""
=================================================
Author : Athena
Time : 2023/8/26 20:07 
=================================================
"""
import json
from openpyxl import load_workbook


def get_test_data_from_excel(file, sheet_name):
    # 1、打开工作薄
    wb = load_workbook(filename=file,read_only=True)
    # 2、获取sheet
    sh = wb[sheet_name]
    row = sh.max_row
    column = sh.max_column
    # 3、读取数据
    data = []
    # 获取第一行拿到所有的key
    keys = []

    # 提取标题行以获取字典的键
    for j in range(1, column + 1):
        keys.append(sh.cell(1, j).value)

    # 遍历行以创建字典
    for i in range(2, row + 1):  # 从第二行开始，假设第一行包含标题
        temp = {}
        for j in range(1, column + 1):
            temp[keys[j - 1]] = sh.cell(i, j).value  # 使用正确的行索引
        try:
            temp['request_data'] = json.loads(temp['request_data'])
            temp['expect_data'] = json.loads(temp['expect_data'])
        except json.decoder.JSONDecodeError:
            raise ValueError('用例数据json格式错误')

        # 把每行数据形成的字典添加到data 列表中
        data.append(temp) 
    return data

if __name__ == '__main__':
    test_data = get_test_data_from_excel('../resource/testcase.xlsx','login')
    print(test_data)